# src/lab05/csv_xlsx.py
import csv
from pathlib import Path

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Использовать openpyxl ИЛИ xlsxwriter.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    """
    try:
        if not Path(csv_path).exists():
            raise FileNotFoundError(f"Файл не найден: {csv_path}")
        
        # Чтение CSV
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
        
        if not rows:
            raise ValueError("CSV файл пуст")
        
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("Требуется установить openpyxl: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Запись данных
        for row_idx, row in enumerate(rows, 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # колонki автоширина
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            ws.column_dimensions[col_letter].width = max(max_length + 2, 8)
        
        # Сохранение
        Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(xlsx_path)
        
    except csv.Error:
        raise ValueError("Ошибка формата CSV")
    except Exception as e:
        if isinstance(e, (FileNotFoundError, ValueError)):
            raise e
        raise ValueError(f"Ошибка конвертации: {e}")